import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font
from io import BytesIO

st.title("Excel Processing App")

# --- Step 1: Upload Excel File ---
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file is not None:
    st.info("Processing your Excel file...")
    wb = load_workbook(uploaded_file)

    # --- Load sheets ---
    AR_Aging_sheet = wb["AR Aging (excluding HUD)"]
    Sample_sheet = wb["Sample Report"]
    Rent_sheet = wb["Rent Roll w. Lease Charges"]
    Legal_sheet = wb["Legal Report"]
    Tenant_memo_sheet = wb["Tenant Memo's"]

    ws_sample = Sample_sheet

    # --- Initialize Sample Sheet Headers ---
    ws_sample.font = Font(bold=True, size=8)
    ws_sample['J2'] = "Unit"
    ws_sample['H2'] = "Name"
    ws_sample['S2'] = "Unit"

    # -------------------------
    # --- LEGAL REPORT DATA ---
    # -------------------------
    unit_legal = []
    status_legal = []
    book_number = []
    legal_type = []
    legal_reason = []
    active_legal = []
    due_date = []
    current_alert = []
    legal_notes = []

    sheet = Legal_sheet
    blank_count = 0
    max_consecutive_blank = 100

    for row in sheet.iter_rows(min_row=7, min_col=2, max_col=18):  # B-R
        unit_cell = row[0]      # B: Unit
        status_cell = row[3]    # E: Status
        f_cell = row[4]         # F: Book Number
        g_cell = row[5]         # G: Legal Type
        h_cell = row[6]         # H: Legal Reason
        i_cell = row[7]         # I: Active Legal
        m_cell = row[11]        # M: Due Date
        n_cell = row[12]        # N: Current Alert
        r_cell = row[16]        # R: Legal Notes

        # Stop after too many blanks
        if all(c.value is None or str(c.value).strip() == "" for c in [unit_cell, status_cell, f_cell, g_cell, h_cell, i_cell, m_cell, n_cell, r_cell]):
            blank_count += 1
            if blank_count > max_consecutive_blank:
                break
            continue
        else:
            blank_count = 0

        def safe_value(cell):
            return cell.value if cell.value is not None and (cell.font is None or not cell.font.bold) else ""

        if unit_cell.font and not unit_cell.font.bold:
            unit_legal.append(str(unit_cell.value).strip())
            status_legal.append(status_cell.value if status_cell.value is not None else "")

        book_number.append(safe_value(f_cell))
        legal_type.append(safe_value(g_cell))
        legal_reason.append(safe_value(h_cell))
        active_legal.append(safe_value(i_cell))
        due_date.append(safe_value(m_cell))
        current_alert.append(safe_value(n_cell))
        legal_notes.append(safe_value(r_cell))

    # Paste Legal data into Sample Sheet
    row_start = 3
    for i, unit in enumerate(unit_legal):
        ws_sample[f"J{row_start + i}"] = unit
        ws_sample[f"K{row_start + i}"] = status_legal[i]
        ws_sample[f"L{row_start + i}"] = book_number[i]
        ws_sample[f"M{row_start + i}"] = legal_type[i]
        ws_sample[f"N{row_start + i}"] = legal_reason[i]
        ws_sample[f"O{row_start + i}"] = active_legal[i]
        ws_sample[f"P{row_start + i}"] = due_date[i]
        ws_sample[f"Q{row_start + i}"] = current_alert[i]
        ws_sample[f"R{row_start + i}"] = legal_notes[i]

    # -------------------------
    # --- AR AGING DATA ---
    # -------------------------
    def process_column(sheet, column, stop_word=None):
        result = []
        for row in range(8, sheet.max_row + 1):
            value = sheet[f"{column}{row}"].value
            if value is None:
                break
            if stop_word and str(value).strip().lower() == stop_word.lower():
                break
            result.append(value)
        return result

    unit_ar = process_column(AR_Aging_sheet, "A", stop_word="total")
    resident_ar = process_column(AR_Aging_sheet, "B")
    status_ar = process_column(AR_Aging_sheet, "C")
    tenant_name_ar = process_column(AR_Aging_sheet, "D")
    total_charges_ar = process_column(AR_Aging_sheet, "E")
    _0_30_ar = process_column(AR_Aging_sheet, "F")

    ar_data = {}
    for u, resident, status, tenant_name, total_charges, ar_0_30 in zip(
        unit_ar, resident_ar, status_ar, tenant_name_ar, total_charges_ar, _0_30_ar):
        ar_data[str(u).strip().upper()] = {
            "resident": resident,
            "status": status,
            "tenant_name": tenant_name,
            "total_charges": total_charges,
            "ar_0_30": ar_0_30
        }

    # Fill Sample Sheet with AR data
    current_row = row_start
    seen_units = set()
    ar_amount_written_units = set()
    for unit in unit_legal:
        key = str(unit).strip().upper()
        if key not in seen_units:
            ws_sample[f"J{current_row}"] = key
            seen_units.add(key)
        else:
            ws_sample[f"J{current_row}"] = ""

        if key in ar_data:
            ws_sample[f"B{current_row}"] = key
            ws_sample[f"C{current_row}"] = ar_data[key]["resident"]
            ws_sample[f"D{current_row}"] = ar_data[key]["status"]
            ws_sample[f"E{current_row}"] = ar_data[key]["tenant_name"]
            if key not in ar_amount_written_units:
                ws_sample[f"F{current_row}"] = ar_data[key]["ar_0_30"]
                ws_sample[f"G{current_row}"] = ar_data[key]["total_charges"]
                ar_amount_written_units.add(key)
        current_row += 1

    # -------------------------
    # --- RENT DATA ---
    # -------------------------
    unit_rent = []
    name_rent = []
    amount_rent = []

    for row in range(8, Rent_sheet.max_row + 1):
        u = Rent_sheet[f"A{row}"].value
        n = Rent_sheet[f"E{row}"].value
        a = Rent_sheet[f"H{row}"].value
        if u is None:
            continue
        if "summary" in str(u).lower():
            break
        unit_rent.append(str(u).strip().upper())
        name_rent.append(n)
        amount_rent.append(a)

    unit_to_name = dict(zip(unit_rent, name_rent))
    unit_to_amount = dict(zip(unit_rent, amount_rent))
    amount_written_units = set()

    current_row = row_start
    for unit in unit_legal:
        key = str(unit).strip().upper()
        if key in unit_to_name:
            ws_sample[f"H{current_row}"] = unit_to_name[key]
        if key in unit_to_amount and key not in amount_written_units:
            ws_sample[f"I{current_row}"] = unit_to_amount[key]
            amount_written_units.add(key)
        current_row += 1

    # -------------------------
    # --- TENANT MEMO DATA ---
    # -------------------------
    tenant_units = []
    type_tenant = []
    date_tenant = []
    memo_tenant = []

    # Units
    blank_count = 0
    for row in range(6, Tenant_memo_sheet.max_row + 1):
        value = Tenant_memo_sheet[f"B{row}"].value
        if value is None or str(value).strip() == "":
            blank_count += 1
            if blank_count > 6:
                break
            continue
        tenant_units.append(value)
        blank_count = 0

    # Type
    blank_count = 0
    for row in range(2, Tenant_memo_sheet.max_row + 1):
        value = Tenant_memo_sheet[f"A{row}"].value
        if value is None or str(value).strip() == "":
            blank_count += 1
            if blank_count > 5:
                break
            continue
        if "type" in str(value).lower():
            type_tenant.append(value)
        blank_count = 0

    # Date
    blank_count = 0
    for row in range(2, Tenant_memo_sheet.max_row + 1):
        value = Tenant_memo_sheet[f"C{row}"].value
        if value is None or str(value).strip() == "":
            blank_count += 1
            if blank_count > 5:
                break
            continue
        if "date" in str(value).lower():
            date_tenant.append(value)
        blank_count = 0

    # Memo
    blank_count = 0
    for row in range(2, Tenant_memo_sheet.max_row + 1):
        cell = Tenant_memo_sheet[f"D{row}"]
        value = cell.value
        if value is None or str(value).strip() == "":
            blank_count += 1
            if blank_count > 5:
                break
            continue
        if cell.font and not cell.font.bold:
            memo_tenant.append(value)
        blank_count = 0

    tenant_data = dict()
    for u, ttype, tdate, tmemo in zip(tenant_units, type_tenant, date_tenant, memo_tenant):
        tenant_data[str(u).strip().upper()] = {"type": ttype, "date": tdate, "memo": tmemo}

    current_row = row_start
    memo_written_units = set()
    for unit in unit_legal:
        key = str(unit).strip().upper()
        if key in tenant_data and key not in memo_written_units:
            ws_sample[f"T{current_row}"] = tenant_data[key]["type"]
            ws_sample[f"U{current_row}"] = tenant_data[key]["date"]
            ws_sample[f"V{current_row}"] = tenant_data[key]["memo"]
            memo_written_units.add(key)
        current_row += 1

    # -------------------------
    # --- SAVE AND DOWNLOAD ---
    # -------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        label="Download Processed Excel",
        data=output,
        file_name="Processed_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.success("Processing complete! Click the button above to download your file.")
